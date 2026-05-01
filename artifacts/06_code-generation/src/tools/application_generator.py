"""
申請書生成ツール

収集済み申請情報とテンプレートを組み合わせて Excel 形式の申請書（下書き）を生成する。
"""
import logging
import os
from datetime import datetime

import openpyxl
from pydantic import ValidationError
from strands import tool
from strands.types.tools import ToolContext

from handlers.error_handler import ErrorHandler
from models.data_models import (
    ExpenseApplicationInput,
    ExpenseItem,
    TransportApplicationInput,
    TransportSegment,
)

logger = logging.getLogger("tools.application_generator")

TRANSPORT_TEMPLATE_PATH = "data/templates/transport_template.xlsx"
EXPENSE_TEMPLATE_PATH = "data/templates/expense_template.xlsx"


def _get_output_path(session_id: str, applicant_name: str, doc_type: str) -> tuple[str, str]:
    """出力ファイルパスとディレクトリを生成する。"""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_dir = f"data/output/{session_id}/"
    file_name = f"{applicant_name}_{doc_type}_{timestamp}.xlsx"
    return output_dir, f"{output_dir}{file_name}"


@tool(context=True)
def generate_transport_application(
    collected_items: dict,
    tool_context: ToolContext,
) -> dict:
    """
    交通費精算申請書生成ツール

    収集済み移動情報リストから交通費精算申請書（下書き）Excel ファイルを生成する。
    申請者名・申請日は invocation_state から取得する（LLM パラメータではない）。
    APR-001（社員の申請書生成前確認）通過後に呼び出すこと。

    Args:
        collected_items: 収集済み移動情報を含む辞書。segments キーに移動区間リストを含む。
        tool_context: invocation_state を含む ToolContext。

    Returns:
        dict: success (bool), file_path (str | None), message (str | None) を含む辞書。
    """
    state = tool_context.invocation_state or {}
    applicant_name = state.get("applicant_name", "")
    application_date = state.get("application_date", "")
    session_id = state.get("session_id", "default")

    logger.info(
        f"交通費精算申請書生成開始: 申請者=**** "
        f"({len(collected_items.get('segments', []))}区間)"
    )

    if "segments" not in collected_items:
        return {
            "success": False,
            "file_path": None,
            "message": "collected_items に 'segments' キーが存在しません。{'segments': [...]} 形式で指定してください。",
        }

    segments_data = collected_items["segments"]
    business_purpose = segments_data[0].get("business_purpose", "業務用") if segments_data else "業務用"
    try:
        validated = TransportApplicationInput(
            business_purpose=business_purpose,
            segments=segments_data,
        )
    except ValidationError as e:
        logger.error(f"交通費精算申請書生成: 入力バリデーションエラー: {str(e)}", exc_info=True)
        return {
            "success": False,
            "file_path": None,
            "message": ErrorHandler.handle_validation_error(e),
        }

    if not os.path.exists(TRANSPORT_TEMPLATE_PATH):
        logger.error(
            f"交通費精算申請書生成: テンプレートファイルが見つかりません: file_path={TRANSPORT_TEMPLATE_PATH}"
        )
        return {
            "success": False,
            "file_path": None,
            "message": ErrorHandler.handle_file_save_error(
                FileNotFoundError(TRANSPORT_TEMPLATE_PATH)
            ),
        }

    try:
        wb = openpyxl.load_workbook(TRANSPORT_TEMPLATE_PATH)
    except Exception as e:
        logger.error(
            f"交通費精算申請書生成: 想定外エラー: file_path={TRANSPORT_TEMPLATE_PATH}",
            exc_info=True,
        )
        return {
            "success": False,
            "file_path": None,
            "message": ErrorHandler.handle_unexpected_error(e),
        }

    ws = wb.active
    ws["B3"] = applicant_name
    ws["B4"] = application_date

    items = validated.segments
    for i, item in enumerate(items):
        row = 7 + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = str(item.travel_date)
        ws[f"C{row}"] = item.departure
        ws[f"D{row}"] = item.destination
        ws[f"E{row}"] = item.transport_type
        ws[f"F{row}"] = item.fare
        ws[f"G{row}"] = item.business_purpose
        ws[f"H{row}"] = ""

    total_row = 7 + len(items) + 2
    ws[f"F{total_row}"] = sum(item.fare for item in items)

    output_dir, file_path = _get_output_path(session_id, applicant_name, "交通費精算申請書")

    try:
        os.makedirs(output_dir, exist_ok=True)
        wb.save(file_path)
    except IOError as e:
        logger.error(f"交通費精算申請書生成: ファイル保存失敗 (IOError): file_path={file_path}", exc_info=True)
        return False, ErrorHandler.handle_file_save_error(e)
    except PermissionError as e:
        logger.error(
            f"交通費精算申請書生成: ファイル保存失敗 (PermissionError): file_path={file_path}",
            exc_info=True,
        )
        return False, ErrorHandler.handle_file_save_error(e)
    except Exception as e:
        logger.error(
            f"交通費精算申請書生成: 想定外エラー: file_path={file_path}", exc_info=True
        )
        return False, ErrorHandler.handle_file_save_error(e)

    logger.info(f"交通費精算申請書生成成功: file_path={file_path}")
    return {"success": True, "file_path": file_path, "message": None}


@tool(context=True)
def generate_expense_application(
    collected_items: dict,
    tool_context: ToolContext,
) -> dict:
    """
    経費精算申請書生成ツール

    収集済み経費明細リストから経費精算申請書（下書き）Excel ファイルを生成する。
    申請者名・申請日は invocation_state から取得する（LLM パラメータではない）。
    APR-001（社員の申請書生成前確認）通過後に呼び出すこと。

    Args:
        collected_items: 収集済み経費情報を含む辞書。expenses キーに経費明細リストを含む。
        tool_context: invocation_state を含む ToolContext。

    Returns:
        dict: success (bool), file_path (str | None), message (str | None) を含む辞書。
    """
    state = tool_context.invocation_state or {}
    applicant_name = state.get("applicant_name", "")
    application_date = state.get("application_date", "")
    session_id = state.get("session_id", "default")

    logger.info(
        f"経費精算申請書生成開始: 申請者=**** "
        f"({len(collected_items.get('expenses', []))}件)"
    )

    if "expenses" not in collected_items:
        return {
            "success": False,
            "file_path": None,
            "message": "collected_items に 'expenses' キーが存在しません。{'expenses': [...]} 形式で指定してください。",
        }

    expenses_data = collected_items["expenses"]
    business_purpose = expenses_data[0].get("business_purpose", "業務用") if expenses_data else "業務用"
    try:
        validated = ExpenseApplicationInput(
            business_purpose=business_purpose,
            expense_items=expenses_data,
        )
    except ValidationError as e:
        logger.error(f"経費精算申請書生成: 入力バリデーションエラー: {str(e)}", exc_info=True)
        return {
            "success": False,
            "file_path": None,
            "message": ErrorHandler.handle_validation_error(e),
        }

    if not os.path.exists(EXPENSE_TEMPLATE_PATH):
        logger.error(
            f"経費精算申請書生成: テンプレートファイルが見つかりません: file_path={EXPENSE_TEMPLATE_PATH}"
        )
        return {
            "success": False,
            "file_path": None,
            "message": ErrorHandler.handle_file_save_error(
                FileNotFoundError(EXPENSE_TEMPLATE_PATH)
            ),
        }

    try:
        wb = openpyxl.load_workbook(EXPENSE_TEMPLATE_PATH)
    except Exception as e:
        logger.error(
            f"経費精算申請書生成: 想定外エラー: file_path={EXPENSE_TEMPLATE_PATH}", exc_info=True
        )
        return {
            "success": False,
            "file_path": None,
            "message": ErrorHandler.handle_unexpected_error(e),
        }

    ws = wb.active
    ws["B3"] = applicant_name
    ws["B4"] = application_date

    items = validated.expense_items
    for i, item in enumerate(items):
        row = 7 + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = str(item.purchase_date)
        ws[f"C{row}"] = item.store_name
        ws[f"D{row}"] = item.item_name
        ws[f"E{row}"] = item.expense_category
        ws[f"F{row}"] = item.amount
        ws[f"G{row}"] = validated.business_purpose
        ws[f"H{row}"] = ""

    total_row = 7 + len(items) + 2
    ws[f"F{total_row}"] = sum(item.amount for item in items)

    output_dir, file_path = _get_output_path(session_id, applicant_name, "経費精算申請書")

    try:
        os.makedirs(output_dir, exist_ok=True)
        wb.save(file_path)
    except IOError as e:
        logger.error(f"経費精算申請書生成: ファイル保存失敗 (IOError): file_path={file_path}", exc_info=True)
        return False, ErrorHandler.handle_file_save_error(e)
    except PermissionError as e:
        logger.error(
            f"経費精算申請書生成: ファイル保存失敗 (PermissionError): file_path={file_path}",
            exc_info=True,
        )
        return False, ErrorHandler.handle_file_save_error(e)
    except Exception as e:
        logger.error(
            f"経費精算申請書生成: 想定外エラー: file_path={file_path}", exc_info=True
        )
        return False, ErrorHandler.handle_file_save_error(e)

    logger.info(f"経費精算申請書生成成功: file_path={file_path}")
    return {"success": True, "file_path": file_path, "message": None}
