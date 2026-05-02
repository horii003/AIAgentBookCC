import logging
import os
import re
from datetime import datetime

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError
from strands import tool
from strands.types.tools import ToolContext

from config.settings import EXPENSE_TEMPLATE_PATH, TRANSPORT_TEMPLATE_PATH
from handlers.error_handler import ErrorHandler
from models.data_models import ExpenseFormInput, TransportFormInput

logger = logging.getLogger(__name__)
error_handler = ErrorHandler()

_TRANSPORT_REQUIRED_KEYS = {"travel_date", "departure", "destination", "transportation_type", "amount", "purpose"}
_EXPENSE_REQUIRED_KEYS = {"store_name", "amount", "expense_date", "item_name", "expense_category", "purpose"}


@tool(context=True)
def generate_transport_expense_form(
    tool_context: ToolContext,
    segments: list,
) -> dict:
    """交通費精算申請書（下書き）を Excel ファイルとして生成する。

    HumanApprovalHook による社員確認（OK）取得後のみ呼び出すこと（GRD-014）。
    申請者名・申請日は invocation_state から取得する。LLM はパラメータとして渡さない。
    出力ファイルパスはツール内部で自動生成する。

    Args:
        segments: 移動情報リスト（1区間以上）。各要素に travel_date, departure, destination,
                  transportation_type, amount, purpose キーが必要。

    Returns:
        dict with keys: success, file_path, message
    """
    applicant_name = tool_context.invocation_state.get("applicant_name", "")
    application_date = tool_context.invocation_state.get("application_date", "")
    session_id = tool_context.invocation_state.get("session_id", "unknown")

    logger.info(f"[OPE-002] generate_transport_expense_form 開始: applicant_name={applicant_name}, segments_count={len(segments) if segments else 0}")

    if not segments:
        return {"success": False, "file_path": "", "message": "移動情報が入力されていません。移動情報を入力してください"}

    for seg in segments:
        missing = _TRANSPORT_REQUIRED_KEYS - set(seg.keys())
        if missing:
            return {
                "success": False,
                "file_path": "",
                "message": "移動情報の必須キーが不足しています。以下のキーをすべて含めてください: travel_date, departure, destination, transportation_type, amount, purpose",
            }

    try:
        form_input = TransportFormInput(
            applicant_name=applicant_name,
            application_date=application_date,
            segments=segments,
        )
    except ValidationError as e:
        return {"success": False, "file_path": "", "message": error_handler.handle_validation_error(e)}

    if not os.path.exists(TRANSPORT_TEMPLATE_PATH):
        logger.error(f"[ERR-005] テンプレートファイル不在: {TRANSPORT_TEMPLATE_PATH}")
        return {"success": False, "file_path": "", "message": "申請書テンプレートが見つかりませんでした。担当部門（管理部）にお問い合わせください。"}

    try:
        wb = openpyxl.load_workbook(TRANSPORT_TEMPLATE_PATH)
    except InvalidFileException as e:
        logger.error(f"[ERR-006] 申請書生成失敗: {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}

    ws = wb.active
    ws["B3"] = form_input.applicant_name
    ws["B4"] = form_input.application_date
    for i, seg in enumerate(form_input.segments):
        row = 7 + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = str(seg.travel_date) if hasattr(seg.travel_date, "isoformat") else seg.travel_date
        ws[f"C{row}"] = seg.departure
        ws[f"D{row}"] = seg.destination
        ws[f"E{row}"] = seg.transportation_type
        ws[f"F{row}"] = seg.amount
        ws[f"G{row}"] = seg.purpose
        ws[f"H{row}"] = ""

    output_dir = os.path.join("data", "output", session_id)
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", form_input.applicant_name)
    file_path = os.path.join(output_dir, f"{safe_name}_交通費精算申請書_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    try:
        wb.save(file_path)
    except IOError as e:
        logger.error(f"[ERR-006] 申請書保存失敗(IOError): {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}
    except PermissionError as e:
        logger.error(f"[ERR-006] 申請書保存失敗(PermissionError): {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}
    except Exception as e:
        logger.error(f"[ERR-006] 申請書生成失敗: {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}

    logger.info(f"[OPE-002] generate_transport_expense_form 完了: file_path={file_path}")
    return {"success": True, "file_path": file_path, "message": ""}


@tool(context=True)
def generate_expense_reimbursement_form(
    tool_context: ToolContext,
    items: list,
) -> dict:
    """経費精算申請書（下書き）を Excel ファイルとして生成する。

    HumanApprovalHook による社員確認（OK）取得後のみ呼び出すこと（GRD-014）。
    申請者名・申請日は invocation_state から取得する。LLM はパラメータとして渡さない。
    出力ファイルパスはツール内部で自動生成する。

    Args:
        items: 経費明細リスト（1明細以上）。各要素に store_name, amount, expense_date,
               item_name, expense_category, purpose キーが必要。

    Returns:
        dict with keys: success, file_path, message
    """
    applicant_name = tool_context.invocation_state.get("applicant_name", "")
    application_date = tool_context.invocation_state.get("application_date", "")
    session_id = tool_context.invocation_state.get("session_id", "unknown")

    logger.info(f"[OPE-002] generate_expense_reimbursement_form 開始: applicant_name={applicant_name}, items_count={len(items) if items else 0}")

    if not items:
        return {"success": False, "file_path": "", "message": "経費情報が入力されていません。経費情報を入力してください"}

    for item in items:
        missing = _EXPENSE_REQUIRED_KEYS - set(item.keys())
        if missing:
            return {
                "success": False,
                "file_path": "",
                "message": "経費情報の必須キーが不足しています。以下のキーをすべて含めてください: store_name, amount, expense_date, item_name, expense_category, purpose",
            }

    try:
        form_input = ExpenseFormInput(
            applicant_name=applicant_name,
            application_date=application_date,
            items=items,
        )
    except ValidationError as e:
        return {"success": False, "file_path": "", "message": error_handler.handle_validation_error(e)}

    if not os.path.exists(EXPENSE_TEMPLATE_PATH):
        logger.error(f"[ERR-005] テンプレートファイル不在: {EXPENSE_TEMPLATE_PATH}")
        return {"success": False, "file_path": "", "message": "申請書テンプレートが見つかりませんでした。担当部門（管理部）にお問い合わせください。"}

    try:
        wb = openpyxl.load_workbook(EXPENSE_TEMPLATE_PATH)
    except InvalidFileException as e:
        logger.error(f"[ERR-006] 申請書生成失敗: {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}

    ws = wb.active
    ws["B3"] = form_input.applicant_name
    ws["B4"] = form_input.application_date
    for i, item in enumerate(form_input.items):
        row = 7 + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = str(item.expense_date) if hasattr(item.expense_date, "isoformat") else item.expense_date
        ws[f"C{row}"] = item.store_name
        ws[f"D{row}"] = item.item_name
        ws[f"E{row}"] = item.expense_category
        ws[f"F{row}"] = item.amount
        ws[f"G{row}"] = item.purpose
        ws[f"H{row}"] = ""

    output_dir = os.path.join("data", "output", session_id)
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", form_input.applicant_name)
    file_path = os.path.join(output_dir, f"{safe_name}_経費精算申請書_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    try:
        wb.save(file_path)
    except IOError as e:
        logger.error(f"[ERR-006] 申請書保存失敗(IOError): {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}
    except PermissionError as e:
        logger.error(f"[ERR-006] 申請書保存失敗(PermissionError): {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}
    except Exception as e:
        logger.error(f"[ERR-006] 申請書生成失敗: {e}")
        return {"success": False, "file_path": "", "message": error_handler.handle_file_save_error(e)}

    logger.info(f"[OPE-002] generate_expense_reimbursement_form 完了: file_path={file_path}")
    return {"success": True, "file_path": file_path, "message": ""}
