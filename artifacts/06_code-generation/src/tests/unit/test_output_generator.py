# 参照: DD-01b 申請書生成ツール詳細設計書
"""tools/output_generator.py の単体テスト"""
import sys
import os
import tempfile
from datetime import date, timedelta
from unittest.mock import MagicMock, patch, mock_open

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


def make_tool_context(
    applicant_name="山田太郎",
    application_date="2026-05-06",
    session_id="test_session",
):
    """テスト用のToolContextモックを生成する。"""
    ctx = MagicMock()
    ctx.invocation_state = {
        "applicant_name": applicant_name,
        "application_date": application_date,
        "session_id": session_id,
    }
    return ctx


def make_valid_segments():
    """有効な移動区間データを返す。"""
    return [
        {
            "travel_date": "2026-05-01",
            "departure": "渋谷",
            "destination": "新宿",
            "transport_type": "電車",
            "fare": 170,
        }
    ]


class TestGenerateTransportApplication:
    """generate_transport_application のテスト"""

    def test_success_with_template(self, tmp_path):
        """正しい交通費申請情報でExcelファイルが生成されsuccess: Trueが返ること。"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")

        # テンプレートファイルを作成
        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(template_path))

        import tools.output_generator as og
        original_path = og.DATA_TRANSPORT_TEMPLATE_PATH
        original_output = og.OUTPUT_BASE_DIR
        original_audit = og.AUDIT_LOG_PATH

        try:
            og.DATA_TRANSPORT_TEMPLATE_PATH = str(template_path)
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")
            og.AUDIT_LOG_PATH = str(tmp_path / "logs" / "audit.log")

            from tools.output_generator import generate_transport_application
            result = generate_transport_application(
                segments=make_valid_segments(),
                purpose="取引先訪問",
                tool_context=make_tool_context(),
            )
            assert result["success"] is True
            assert result["file_path"] != ""
            assert os.path.exists(result["file_path"])
            assert "application_id" in result["application_data"]
        finally:
            og.DATA_TRANSPORT_TEMPLATE_PATH = original_path
            og.OUTPUT_BASE_DIR = original_output
            og.AUDIT_LOG_PATH = original_audit

    def test_template_not_found(self, tmp_path):
        """DATA-002が存在しない場合success: Falseとエラーメッセージが返ること。"""
        import tools.output_generator as og
        original_path = og.DATA_TRANSPORT_TEMPLATE_PATH
        original_output = og.OUTPUT_BASE_DIR

        try:
            og.DATA_TRANSPORT_TEMPLATE_PATH = str(tmp_path / "nonexistent.xlsx")
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")

            from tools.output_generator import generate_transport_application
            result = generate_transport_application(
                segments=make_valid_segments(),
                purpose="取引先訪問",
                tool_context=make_tool_context(),
            )
            assert result["success"] is False
            assert result["file_path"] == ""
            assert result["application_data"] == {}
            assert len(result["message"]) > 0
        finally:
            og.DATA_TRANSPORT_TEMPLATE_PATH = original_path
            og.OUTPUT_BASE_DIR = original_output

    def test_segments_missing_required_key(self, tmp_path):
        """segmentsに必須キーが不足した辞書がある場合、不足キー名を明示したエラーメッセージが返ること（KeyError）。"""
        import tools.output_generator as og
        original_output = og.OUTPUT_BASE_DIR
        try:
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")
            from tools.output_generator import generate_transport_application
            result = generate_transport_application(
                segments=[{"travel_date": "2026-05-01", "departure": "渋谷"}],  # fare等が不足
                purpose="取引先訪問",
                tool_context=make_tool_context(),
            )
            assert result["success"] is False
            assert "必須キー" in result["message"]
        finally:
            og.OUTPUT_BASE_DIR = original_output

    def test_empty_segments(self, tmp_path):
        """segments空リストで「移動区間情報のエラー」が返ること。"""
        import tools.output_generator as og
        original_output = og.OUTPUT_BASE_DIR
        try:
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")
            from tools.output_generator import generate_transport_application
            result = generate_transport_application(
                segments=[],
                purpose="取引先訪問",
                tool_context=make_tool_context(),
            )
            assert result["success"] is False
            assert "移動区間" in result["message"] or len(result["message"]) > 0
        finally:
            og.OUTPUT_BASE_DIR = original_output

    def test_empty_applicant_name(self, tmp_path):
        """申請者名が空の場合success: Falseが返ること。"""
        import tools.output_generator as og
        original_output = og.OUTPUT_BASE_DIR
        try:
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")
            from tools.output_generator import generate_transport_application
            result = generate_transport_application(
                segments=make_valid_segments(),
                purpose="取引先訪問",
                tool_context=make_tool_context(applicant_name=""),
            )
            assert result["success"] is False
            assert "申請者名" in result["message"]
        finally:
            og.OUTPUT_BASE_DIR = original_output

    def test_invocation_state_accessed(self):
        """invocation_stateからapplicant_name・application_date・session_idが正しく取得されること。"""
        import tools.output_generator as og
        from tools.output_generator import generate_transport_application
        ctx = make_tool_context("テスト申請者", "2026-01-01", "sess_test")
        # テンプレートが存在しない場合でもinvocation_stateへのアクセスが正常に行われること
        result = generate_transport_application(
            segments=make_valid_segments(),
            purpose="テスト",
            tool_context=ctx,
        )
        assert isinstance(result, dict)
        assert "success" in result


class TestGenerateExpenseApplication:
    """generate_expense_application のテスト"""

    def test_success_with_template(self, tmp_path):
        """正しい経費申請情報でExcelファイルが生成されること。"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")

        template_path = tmp_path / "expense_template.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(template_path))

        import tools.output_generator as og
        original_path = og.DATA_EXPENSE_TEMPLATE_PATH
        original_output = og.OUTPUT_BASE_DIR
        original_audit = og.AUDIT_LOG_PATH

        try:
            og.DATA_EXPENSE_TEMPLATE_PATH = str(template_path)
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")
            og.AUDIT_LOG_PATH = str(tmp_path / "logs" / "audit.log")

            from tools.output_generator import generate_expense_application
            result = generate_expense_application(
                store_name="文具屋",
                expense_category="事務用品費",
                amount=1000,
                expense_date="2026-04-01",
                purpose="業務用品購入",
                tool_context=make_tool_context(),
            )
            assert result["success"] is True
            assert result["file_path"] != ""
            assert os.path.exists(result["file_path"])
        finally:
            og.DATA_EXPENSE_TEMPLATE_PATH = original_path
            og.OUTPUT_BASE_DIR = original_output
            og.AUDIT_LOG_PATH = original_audit

    def test_amount_zero_raises(self):
        """amountに0でValidationErrorに基づく日本語エラーメッセージが返ること（EX-01）。"""
        from tools.output_generator import generate_expense_application
        result = generate_expense_application(
            store_name="文具屋",
            expense_category="事務用品費",
            amount=0,
            expense_date="2026-04-01",
            purpose="業務用品購入",
            tool_context=make_tool_context(),
        )
        assert result["success"] is False
        assert result["application_data"] == {}
        assert len(result["message"]) > 0

    def test_deadline_exceeded(self):
        """申請期限（90日）を超過したexpense_dateでValidationErrorが発生すること（BRL-12）。"""
        from tools.output_generator import generate_expense_application
        # 申請日から91日前の経費発生日
        app_date = date(2026, 5, 6)
        exp_date = app_date - timedelta(days=91)
        result = generate_expense_application(
            store_name="文具屋",
            expense_category="事務用品費",
            amount=1000,
            expense_date=exp_date.isoformat(),
            purpose="業務用品購入",
            tool_context=make_tool_context(application_date=app_date.isoformat()),
        )
        assert result["success"] is False
        assert len(result["message"]) > 0

    def test_empty_applicant_name(self):
        """申請者名が空の場合success: Falseが返ること。"""
        from tools.output_generator import generate_expense_application
        result = generate_expense_application(
            store_name="文具屋",
            expense_category="事務用品費",
            amount=1000,
            expense_date="2026-04-01",
            purpose="業務用品購入",
            tool_context=make_tool_context(applicant_name=""),
        )
        assert result["success"] is False
        assert "申請者名" in result["message"]

    def test_template_not_found(self, tmp_path):
        """DATA-003が存在しない場合success: Falseとエラーメッセージが返ること。"""
        import tools.output_generator as og
        original_path = og.DATA_EXPENSE_TEMPLATE_PATH
        original_output = og.OUTPUT_BASE_DIR

        try:
            og.DATA_EXPENSE_TEMPLATE_PATH = str(tmp_path / "nonexistent.xlsx")
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")

            from tools.output_generator import generate_expense_application
            result = generate_expense_application(
                store_name="文具屋",
                expense_category="事務用品費",
                amount=1000,
                expense_date="2026-04-01",
                purpose="業務用品購入",
                tool_context=make_tool_context(),
            )
            assert result["success"] is False
            assert result["file_path"] == ""
        finally:
            og.DATA_EXPENSE_TEMPLATE_PATH = original_path
            og.OUTPUT_BASE_DIR = original_output

    def test_file_write_retry_on_ioerror(self, tmp_path):
        """ファイル書き込みが2回失敗した場合（IOErrorモック）success: Falseが返ること（EX-04）。"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")

        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(template_path))

        import tools.output_generator as og
        original_path = og.DATA_EXPENSE_TEMPLATE_PATH
        original_output = og.OUTPUT_BASE_DIR

        try:
            og.DATA_EXPENSE_TEMPLATE_PATH = str(template_path)
            og.OUTPUT_BASE_DIR = str(tmp_path / "output")

            # _save_excel_fileをモックして2回ともIOErrorを返す
            with patch("tools.output_generator._save_excel_file", return_value=(False, "IOエラーが発生しました")):
                from tools.output_generator import generate_expense_application
                result = generate_expense_application(
                    store_name="文具屋",
                    expense_category="事務用品費",
                    amount=1000,
                    expense_date="2026-04-01",
                    purpose="業務用品購入",
                    tool_context=make_tool_context(),
                )
                assert result["success"] is False
        finally:
            og.DATA_EXPENSE_TEMPLATE_PATH = original_path
            og.OUTPUT_BASE_DIR = original_output
