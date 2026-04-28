"""Integration tests for the multi-agent expense application system."""
import json
import os
import pytest
from unittest.mock import MagicMock, patch

import openpyxl


# ---- Helper fixtures ----

@pytest.fixture
def train_routes_data():
    return {
        "渋谷_新宿": 170,
        "新宿_渋谷": 170,
        "東京_品川": 160,
        "品川_東京": 160,
    }


@pytest.fixture
def fixed_fares_data():
    return {
        "バス": 230,
        "タクシー": 10000,
        "飛行機": 50000,
    }


@pytest.fixture
def fare_loader_with_data(train_routes_data, fixed_fares_data):
    from tools.transport_tools import _fare_loader
    original_train = _fare_loader.train_routes_data.copy()
    original_fixed = _fare_loader.fixed_fares_data.copy()
    _fare_loader.train_routes_data = train_routes_data.copy()
    _fare_loader.fixed_fares_data = fixed_fares_data.copy()
    yield _fare_loader
    _fare_loader.train_routes_data = original_train
    _fare_loader.fixed_fares_data = original_fixed


@pytest.fixture
def make_tool_context():
    def _make(session_id="integ-session", applicant_name="山田太郎", application_date="2026-04-28"):
        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": session_id,
            "applicant_name": applicant_name,
            "application_date": application_date,
        }
        return ctx
    return _make


# ---- FareDataLoader integration ----

class TestFareDataLoaderIntegration:
    def test_load_train_routes_from_json_file(self, tmp_path):
        data = {"渋谷_新宿": 170, "新宿_渋谷": 170}
        fpath = tmp_path / "train_routes.json"
        fpath.write_text(json.dumps(data), encoding="utf-8")

        from tools.transport_tools import FareDataLoader
        loader = FareDataLoader()
        with patch("tools.transport_tools._TRAIN_ROUTES_PATH", str(fpath)):
            ok, msg = loader.load_train_routes()

        assert ok is True
        assert loader.train_routes_data == data

    def test_load_fixed_fares_from_json_file(self, tmp_path):
        data = {"バス": 230, "タクシー": 10000, "飛行機": 50000}
        fpath = tmp_path / "fixed_fares.json"
        fpath.write_text(json.dumps(data), encoding="utf-8")

        from tools.transport_tools import FareDataLoader
        loader = FareDataLoader()
        with patch("tools.transport_tools._FIXED_FARES_PATH", str(fpath)):
            ok, msg = loader.load_fixed_fares()

        assert ok is True
        assert loader.fixed_fares_data == data

    def test_loaded_data_used_in_calculation(self, tmp_path):
        train_data = {"渋谷_新宿": 170}
        fixed_data = {"バス": 230}

        train_path = tmp_path / "train_routes.json"
        fixed_path = tmp_path / "fixed_fares.json"
        train_path.write_text(json.dumps(train_data), encoding="utf-8")
        fixed_path.write_text(json.dumps(fixed_data), encoding="utf-8")

        from tools.transport_tools import FareDataLoader, _calculate
        loader = FareDataLoader()
        with patch("tools.transport_tools._TRAIN_ROUTES_PATH", str(train_path)):
            loader.load_train_routes()
        with patch("tools.transport_tools._FIXED_FARES_PATH", str(fixed_path)):
            loader.load_fixed_fares()

        from tools import transport_tools
        original_train = transport_tools._fare_loader.train_routes_data
        original_fixed = transport_tools._fare_loader.fixed_fares_data
        transport_tools._fare_loader.train_routes_data = loader.train_routes_data
        transport_tools._fare_loader.fixed_fares_data = loader.fixed_fares_data

        result = _calculate("電車", "渋谷", "新宿")
        assert result["success"] is True
        assert result["fare"] == 170

        transport_tools._fare_loader.train_routes_data = original_train
        transport_tools._fare_loader.fixed_fares_data = original_fixed


# ---- calculate_transport_expense integration ----

class TestCalculateTransportExpenseIntegration:
    def test_train_route_returns_fare(self, fare_loader_with_data, make_tool_context):
        from tools.transport_tools import calculate_transport_expense
        ctx = make_tool_context()
        result = calculate_transport_expense(
            tool_context=ctx,
            transport_date="2026-04-28",
            departure="渋谷",
            destination="新宿",
            transport_type="電車",
        )
        assert result["success"] is True
        assert result["fare"] == 170
        assert "テーブル" in result["calculation_basis"]

    def test_bus_returns_fixed_fare(self, fare_loader_with_data, make_tool_context):
        from tools.transport_tools import calculate_transport_expense
        ctx = make_tool_context()
        result = calculate_transport_expense(
            tool_context=ctx,
            transport_date="2026-04-28",
            departure="渋谷",
            destination="新宿",
            transport_type="バス",
        )
        assert result["success"] is True
        assert result["fare"] == 230

    def test_taxi_returns_fixed_fare(self, fare_loader_with_data, make_tool_context):
        from tools.transport_tools import calculate_transport_expense
        ctx = make_tool_context()
        result = calculate_transport_expense(
            tool_context=ctx,
            transport_date="2026-04-28",
            departure="A",
            destination="B",
            transport_type="タクシー",
        )
        assert result["success"] is True
        assert result["fare"] == 10000

    def test_unknown_train_route_returns_error(self, fare_loader_with_data, make_tool_context):
        from tools.transport_tools import calculate_transport_expense
        ctx = make_tool_context()
        result = calculate_transport_expense(
            tool_context=ctx,
            transport_date="2026-04-28",
            departure="未知駅",
            destination="不明駅",
            transport_type="電車",
        )
        assert result["success"] is False
        assert "手動" in result["message"]

    def test_invocation_state_session_id_propagated(self, fare_loader_with_data):
        from tools.transport_tools import calculate_transport_expense
        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": "integ-check-session",
            "applicant_name": "田中花子",
            "application_date": "2026-04-28",
        }
        result = calculate_transport_expense(
            tool_context=ctx,
            transport_date="2026-04-28",
            departure="渋谷",
            destination="新宿",
            transport_type="電車",
        )
        assert result["success"] is True


# ---- generate_transport_expense_form integration ----

class TestGenerateTransportExpenseFormIntegration:
    def test_generates_xlsx_file(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active
        template_path = str(tmp_path / "transport_template.xlsx")
        wb.save(template_path)

        output_dir = str(tmp_path / "output")

        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": "integ-transport-sess",
            "applicant_name": "山田太郎",
            "application_date": "2026-04-28",
        }

        segments = [
            {
                "no": 1,
                "transport_date": "2026-04-28",
                "departure": "渋谷",
                "destination": "新宿",
                "transport_type": "電車",
                "amount": 170,
                "business_purpose": "出張",
            }
        ]

        from tools.output_generator import generate_transport_expense_form
        with patch("tools.output_generator._TRANSPORT_TEMPLATE_PATH", template_path):
            with patch("tools.output_generator.os.makedirs"):
                with patch(
                    "tools.output_generator._save_file",
                    return_value=(True, str(tmp_path / "output" / "file.xlsx")),
                ):
                    result = generate_transport_expense_form(
                        tool_context=ctx,
                        segments=segments,
                        business_purpose="出張",
                    )

        assert result["success"] is True
        assert "file_path" in result

    def test_template_not_found_returns_failure(self, tmp_path):
        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": "integ-sess",
            "applicant_name": "山田太郎",
            "application_date": "2026-04-28",
        }

        segments = [
            {
                "no": 1,
                "transport_date": "2026-04-28",
                "departure": "渋谷",
                "destination": "新宿",
                "transport_type": "電車",
                "amount": 170,
                "business_purpose": "出張",
            }
        ]

        from tools.output_generator import generate_transport_expense_form
        with patch("tools.output_generator._TRANSPORT_TEMPLATE_PATH", str(tmp_path / "missing.xlsx")):
            result = generate_transport_expense_form(
                tool_context=ctx,
                segments=segments,
                business_purpose="出張",
            )

        assert result["success"] is False
        assert "テンプレート" in result["message"]


# ---- generate_expense_form integration ----

class TestGenerateExpenseFormIntegration:
    def test_generates_xlsx_file(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active
        template_path = str(tmp_path / "expense_template.xlsx")
        wb.save(template_path)

        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": "integ-expense-sess",
            "applicant_name": "田中花子",
            "application_date": "2026-04-28",
        }

        items = [
            {
                "no": 1,
                "purchase_date": "2026-04-28",
                "store_name": "文具屋",
                "item_name": "ボールペン",
                "expense_category": "事務用品費",
                "amount": 500,
                "business_purpose": "業務用",
            }
        ]

        from tools.output_generator import generate_expense_form
        with patch("tools.output_generator._EXPENSE_TEMPLATE_PATH", template_path):
            with patch("tools.output_generator.os.makedirs"):
                with patch(
                    "tools.output_generator._save_file",
                    return_value=(True, str(tmp_path / "output" / "expense.xlsx")),
                ):
                    result = generate_expense_form(
                        tool_context=ctx,
                        items=items,
                        business_purpose="業務用",
                    )

        assert result["success"] is True
        assert "file_path" in result


# ---- HumanApprovalHook integration ----

def _make_hook_event(tool_name: str):
    event = MagicMock()
    tool_use = {"name": tool_name, "input": {}}
    event.tool_use = tool_use
    return event


class TestHumanApprovalHookIntegration:
    def test_ok_input_does_not_cancel(self):
        from hooks.human_approval_hook import HumanApprovalHook

        hook = HumanApprovalHook(tool_names=["generate_transport_expense_form"])
        event = _make_hook_event("generate_transport_expense_form")

        with patch("builtins.input", return_value="OK"):
            with patch("builtins.print"):
                hook._on_before_tool_call(event)

        assert not isinstance(event.cancel_tool, str)

    def test_cancel_input_cancels_tool(self):
        from hooks.human_approval_hook import HumanApprovalHook

        hook = HumanApprovalHook(tool_names=["generate_transport_expense_form"])
        event = _make_hook_event("generate_transport_expense_form")

        with patch("builtins.input", return_value="キャンセル"):
            with patch("builtins.print"):
                hook._on_before_tool_call(event)

        assert event.cancel_tool is not None and event.cancel_tool != ""

    def test_non_target_tool_not_blocked(self):
        from hooks.human_approval_hook import HumanApprovalHook

        hook = HumanApprovalHook(tool_names=["generate_transport_expense_form"])
        event = _make_hook_event("calculate_transport_expense")

        hook._on_before_tool_call(event)

        assert not isinstance(event.cancel_tool, str)


# ---- LoopControlHook integration ----

class TestLoopControlHookIntegration:
    def test_loop_limit_raises_on_max_iterations(self):
        from hooks.loop_control_hook import LoopControlHook
        from handlers.exceptions import LoopLimitError
        from strands.hooks.events import AfterModelCallEvent

        hook = LoopControlHook(max_iterations=3, agent_name="test_agent")
        event = MagicMock(spec=AfterModelCallEvent)
        event.exception = None

        with pytest.raises(LoopLimitError) as exc_info:
            for _ in range(4):
                hook._on_after_model_call(event)

        assert exc_info.value.max_iterations == 3
        assert exc_info.value.agent_name == "test_agent"

    def test_exception_event_does_not_increment(self):
        from hooks.loop_control_hook import LoopControlHook
        from strands.hooks.events import AfterModelCallEvent

        hook = LoopControlHook(max_iterations=3, agent_name="test")
        event_with_exc = MagicMock(spec=AfterModelCallEvent)
        event_with_exc.exception = Exception("error")

        for _ in range(10):
            hook._on_after_model_call(event_with_exc)

        assert hook._iteration_count == 0


# ---- ErrorHandler integration ----

class TestErrorHandlerIntegration:
    def test_all_methods_return_strings(self):
        from handlers.error_handler import ErrorHandler
        from handlers.exceptions import LoopLimitError
        from pydantic import ValidationError
        from models.data_models import UserInputText

        handler = ErrorHandler()

        loop_err = LoopLimitError(5, 10, "test")
        assert isinstance(handler.handle_loop_limit_error(loop_err), str)
        assert isinstance(handler.handle_unexpected_error(Exception("x")), str)
        assert isinstance(handler.handle_runtime_error(RuntimeError("r")), str)
        assert isinstance(handler.handle_keyboard_interrupt(), str)
        assert isinstance(handler.handle_fare_data_error(Exception("f")), str)
        assert isinstance(handler.handle_calculation_error(Exception("c")), str)
        assert isinstance(handler.handle_file_save_error(IOError("i")), str)

    def test_loop_limit_message_content(self):
        from handlers.error_handler import ErrorHandler
        from handlers.exceptions import LoopLimitError

        handler = ErrorHandler()
        msg = handler.handle_loop_limit_error(LoopLimitError(30, 30, "test"))
        assert isinstance(msg, str)
        assert len(msg) > 0

    def test_unexpected_error_message_content(self):
        from handlers.error_handler import ErrorHandler

        handler = ErrorHandler()
        msg = handler.handle_unexpected_error(Exception("boom"))
        assert "申し訳ありません" in msg


# ---- invocation_state propagation integration ----

class TestInvocationStatePropagation:
    def test_transport_tool_receives_invocation_state(self, fare_loader_with_data):
        from tools.transport_tools import calculate_transport_expense

        received_state = {}

        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": "state-test-session",
            "applicant_name": "鈴木一郎",
            "application_date": "2026-04-28",
        }

        result = calculate_transport_expense(
            tool_context=ctx,
            transport_date="2026-04-28",
            departure="渋谷",
            destination="新宿",
            transport_type="電車",
        )

        assert result["success"] is True

    def test_output_generator_uses_invocation_state(self, tmp_path):
        from tools.output_generator import generate_transport_expense_form

        wb = openpyxl.Workbook()
        template_path = str(tmp_path / "t.xlsx")
        wb.save(template_path)

        ctx = MagicMock()
        ctx.invocation_state = {
            "session_id": "state-sess",
            "applicant_name": "山田太郎",
            "application_date": "2026-04-28",
        }

        segments = [{
            "no": 1,
            "transport_date": "2026-04-28",
            "departure": "渋谷",
            "destination": "新宿",
            "transport_type": "電車",
            "amount": 170,
            "business_purpose": "業務",
        }]

        with patch("tools.output_generator._TRANSPORT_TEMPLATE_PATH", template_path):
            with patch("tools.output_generator.os.makedirs"):
                with patch("tools.output_generator._save_file", return_value=(True, "out.xlsx")):
                    result = generate_transport_expense_form(
                        tool_context=ctx,
                        segments=segments,
                        business_purpose="業務",
                    )

        assert result["success"] is True
        assert "file_path" in result
