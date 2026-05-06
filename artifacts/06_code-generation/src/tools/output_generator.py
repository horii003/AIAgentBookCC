# 参照: DD-01b 申請書生成ツール詳細設計書
"""申請書生成ツール定義

収集済み申請情報（D-005）を業務固有フォーマット（Excel）の申請書に変換してファイルを生成する。
申請者名・申請日はinvocation_stateから取得する（LLMからパラメータとして渡されない）。
生成した申請書を data/output/{session_id}/ ディレクトリに保存する。
"""
import os
import json
import logging
from datetime import datetime
from pathlib import Path

from pydantic import ValidationError

from handlers.error_handler import ErrorHandler
from models.data_models import TransportApplicationData, TransportSegment, ExpenseApplicationData

logger = logging.getLogger(__name__)
_error_handler = ErrorHandler()

# DD-01b 9.1節: テンプレートファイルパス定数
DATA_TRANSPORT_TEMPLATE_PATH = "data/templates/交通費精算申請書テンプレート.xlsx"
DATA_EXPENSE_TEMPLATE_PATH = "data/templates/経費精算申請書テンプレート.xlsx"
OUTPUT_BASE_DIR = "data/output"

# 監査ログ出力先
AUDIT_LOG_PATH = "logs/audit.log"


def _ensure_directory(path: str) -> None:
    """ディレクトリが存在しない場合は作成する。"""
    Path(path).mkdir(parents=True, exist_ok=True)


def _record_audit_log(
    application_id: str,
    applicant_name: str,
    application_type: str,
    file_path: str,
) -> None:
    """監査ログ（DATA-008）に申請書生成記録を追記する。

    Args:
        application_id: 申請書ID
        applicant_name: 申請者名
        application_type: 申請種別
        file_path: 生成ファイルパス
    """
    try:
        _ensure_directory(os.path.dirname(AUDIT_LOG_PATH) or "logs")
        record = {
            "application_id": application_id,
            "generated_at": datetime.now().isoformat(),
            "applicant_name": applicant_name,
            "application_type": application_type,
            "file_path": file_path,
        }
        with open(AUDIT_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception as e:
        logger.warning("監査ログの記録に失敗しました: %s", e)


def _save_excel_file(workbook, file_path: str) -> tuple:
    """Excelファイルを保存する（最大2回リトライ）。

    Args:
        workbook: openpyxlワークブックオブジェクト
        file_path: 保存先ファイルパス

    Returns:
        tuple:
            成功: (True, "")
            失敗: (False, エラーメッセージ文字列)
    """
    for attempt in range(1, 3):
        try:
            workbook.save(file_path)
            return (True, "")
        except PermissionError as e:
            if attempt == 1:
                logger.warning(
                    "generate file write failed (attempt 1): %s, file_path=%s", e, file_path
                )
            else:
                logger.error(
                    "generate file write failed (attempt 2): %s, file_path=%s", e, file_path
                )
                return (False, _error_handler.handle_file_save_error(e))
        except IOError as e:
            if attempt == 1:
                logger.warning(
                    "generate file write failed (attempt 1): %s, file_path=%s", e, file_path
                )
            else:
                logger.error(
                    "generate file write failed (attempt 2): %s, file_path=%s", e, file_path
                )
                return (False, _error_handler.handle_file_save_error(e))
        except Exception as e:
            if attempt == 1:
                logger.warning(
                    "generate file write failed (attempt 1): %s, file_path=%s", e, file_path
                )
            else:
                logger.error(
                    "generate file write failed (attempt 2): %s, file_path=%s", e, file_path
                )
                return (False, _error_handler.handle_file_save_error(e))
    return (False, "申請書ファイルの保存に失敗しました。")


try:
    from strands import tool, ToolContext
    _STRANDS_AVAILABLE = True
except ImportError:
    _STRANDS_AVAILABLE = False

    def tool(context=False):
        def decorator(func):
            return func
        return decorator if context else lambda func: func

    class ToolContext:
        def __init__(self):
            self.invocation_state = {}


try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    openpyxl = None


@tool(context=True)
def generate_transport_application(
    segments: list,
    purpose: str,
    tool_context: "ToolContext" = None,
) -> dict:
    """収集済みの交通費精算申請情報からExcel申請書を生成してファイルに出力します。

    申請書IDおよびファイルパスはツール内部で自動生成します。
    申請者名・申請日はinvocation_stateから取得します（LLMはこれらをパラメータとして渡しません）。
    ファイル書き込みは最大2回リトライします。

    Args:
        segments (list[dict]): 移動区間情報リスト。各要素は以下のキーを持つ辞書:
            - travel_date (str): 移動日（YYYY-MM-DD形式）
            - departure (str): 出発地
            - destination (str): 目的地
            - transport_type (str): 交通手段（「電車」「バス」「タクシー」「飛行機」）
            - fare (int): 運賃（1以上の正整数）
        purpose (str): 業務目的（空文字禁止・500文字以内）
        tool_context: ツールコンテキスト（invocation_stateを含む）

    Returns:
        dict: 以下のキーを持つ辞書:
            - success (bool): 生成成否フラグ
            - file_path (str): 生成された申請書のファイルパス。エラー時は空文字列
            - application_data (dict): 申請書の各項目と記入値。エラー時は空辞書
            - message (str): エラーメッセージ（success=Trueの場合は空文字列）
    """
    # invocation_stateからapplicant_name・application_date・session_idを取得する
    applicant_name = ""
    application_date = ""
    session_id = ""
    file_path = ""

    if tool_context and hasattr(tool_context, "invocation_state") and tool_context.invocation_state:
        applicant_name = tool_context.invocation_state.get("applicant_name", "")
        application_date = tool_context.invocation_state.get("application_date", "")
        session_id = tool_context.invocation_state.get("session_id", "")

    logger.info(
        "generate_transport_application called: applicant=%s, date=%s",
        applicant_name, application_date,
    )

    try:
        # DD-01b 3.3.2節: segments必須キーガード
        required_segment_keys = ["travel_date", "departure", "destination", "transport_type", "fare"]
        if segments:
            for seg in segments:
                for key in required_segment_keys:
                    if key not in seg:
                        raise KeyError(
                            f"移動区間情報のエラー: 必須キー「{key}」が含まれていません。"
                            f"以下の全キーを含めて再度入力してください: "
                            + ", ".join(required_segment_keys)
                        )

        # TransportApplicationDataでバリデーション（applicant_name・application_dateはinvocation_stateから）
        validated = TransportApplicationData(
            applicant_name=applicant_name if applicant_name else "（未設定）",
            application_date=application_date if application_date else "2000-01-01",
            segments=segments,
            purpose=purpose,
        )

        # 申請者名の空文字チェック
        if not applicant_name or not applicant_name.strip():
            logger.error(
                "generate_transport_application validation error: applicant_name is empty, file_path=%s",
                file_path,
            )
            return {
                "success": False,
                "file_path": "",
                "application_data": {},
                "message": "入力のエラー: 申請者名が設定されていません。",
            }

        if not application_date or not application_date.strip():
            return {
                "success": False,
                "file_path": "",
                "application_data": {},
                "message": "日付のエラー: 申請日が正しく設定されていません。",
            }

    except (ValidationError, KeyError) as e:
        logger.warning(
            "generate_transport_application validation error: %s, file_path=%s", e, file_path
        )
        error_message = _error_handler.handle_validation_error(e)
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": error_message,
        }
    except Exception as e:
        logger.error(
            "generate_transport_application unexpected error: %s, file_path=%s", e, file_path,
            exc_info=True,
        )
        error_message = _error_handler.handle_unexpected_error(e)
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": error_message,
        }

    # DD-01b 3.3.1節: 申請書IDとファイルパスを自律生成
    application_id = f"APP_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    logger.info("application_id generated: %s", application_id)

    file_name = f"{application_id}_交通費精算申請.xlsx"
    output_dir = os.path.join(OUTPUT_BASE_DIR, session_id) if session_id else OUTPUT_BASE_DIR
    file_path = os.path.join(output_dir, file_name)

    # data/output/{session_id}/ ディレクトリを作成する
    try:
        _ensure_directory(output_dir)
    except Exception as e:
        logger.error(
            "generate_transport_application dir creation failed: %s, file_path=%s", e, file_path,
            exc_info=True,
        )
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": "出力ディレクトリの作成に失敗しました。担当部署へご連絡ください。",
        }

    # テンプレートファイルの存在チェック（os.path.exists()による事前チェック）
    if not os.path.exists(DATA_TRANSPORT_TEMPLATE_PATH):
        logger.error(
            "generate_transport_application template not found: %s, file_path=%s",
            DATA_TRANSPORT_TEMPLATE_PATH, file_path,
        )
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": "申請書テンプレートの読み込みに失敗しました。担当部署へご連絡ください。",
        }

    # openpyxlでテンプレート読み込み・データマッピング
    try:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed")

        wb = openpyxl.load_workbook(DATA_TRANSPORT_TEMPLATE_PATH)
        ws = wb.active

        # DD-01b 5.4節: 交通費精算申請書
        # テンプレート構造: 1行目=タイトル, 3行目=申請者名ラベル(A3)/値(B3),
        #   4行目=申請日ラベル(A4)/値(B4), 6行目=ヘッダー, 7行目〜=明細, 9行目G=合計ラベル/H=合計値
        ws["B3"] = applicant_name
        ws["B4"] = application_date

        total_fare = sum(seg.fare for seg in validated.segments)
        start_row = 7  # ヘッダー行(6)の次から明細を記入
        for i, seg in enumerate(validated.segments):
            row = start_row + i
            ws.cell(row=row, column=1, value=i + 1)           # A: No
            ws.cell(row=row, column=2, value=str(seg.travel_date))  # B: 移動日
            ws.cell(row=row, column=3, value=seg.departure)    # C: 出発地
            ws.cell(row=row, column=4, value=seg.destination)  # D: 目的地
            ws.cell(row=row, column=5, value=seg.transport_type)  # E: 交通手段
            ws.cell(row=row, column=6, value=seg.fare)         # F: 費用
            ws.cell(row=row, column=7, value=validated.purpose)  # G: 業務目的
            ws.cell(row=row, column=8, value="")               # H: 承認状況（空欄）

        ws["H9"] = total_fare  # 合計交通費

    except Exception as e:
        logger.error(
            "generate_transport_application template load error: %s, file_path=%s", e, file_path,
            exc_info=True,
        )
        error_message = _error_handler.handle_unexpected_error(e)
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": error_message,
        }

    # application_dataの構築（CF-005チェック参照用）
    application_data = {
        "application_id": application_id,
        "applicant_name": applicant_name,
        "application_date": application_date,
        "segments": [
            {
                "travel_date": str(seg.travel_date),
                "departure": seg.departure,
                "destination": seg.destination,
                "transport_type": seg.transport_type,
                "fare": seg.fare,
            }
            for seg in validated.segments
        ],
        "purpose": validated.purpose,
        "total_fare": sum(seg.fare for seg in validated.segments),
    }

    # Excelファイルを保存する（最大2回リトライ）
    save_result, save_error = _save_excel_file(wb, file_path)
    if not save_result:
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": save_error,
        }

    # 監査ログ（DATA-008）に記録する
    _record_audit_log(
        application_id=application_id,
        applicant_name=applicant_name,
        application_type="交通費精算申請",
        file_path=file_path,
    )

    logger.info("generate_transport_application succeeded: file_path=%s", file_path)
    return {
        "success": True,
        "file_path": file_path,
        "application_data": application_data,
        "message": "",
    }


@tool(context=True)
def generate_expense_application(
    store_name: str,
    expense_category: str,
    amount: int,
    expense_date: str,
    purpose: str,
    tool_context: "ToolContext" = None,
) -> dict:
    """収集済みの経費精算申請情報からExcel申請書を生成してファイルに出力します。

    申請書IDおよびファイルパスはツール内部で自動生成します。
    申請者名・申請日はinvocation_stateから取得します（LLMはこれらをパラメータとして渡しません）。
    ファイル書き込みは最大2回リトライします。

    Args:
        store_name (str): 店舗名（空文字禁止・500文字以内）
        expense_category (str): 経費区分。以下の表記を許容し内部統一表記へ正規化します:
            - 「事務用品費」「事務用品」→「事務用品費」
            - 「宿泊費」「宿泊」→「宿泊費」
            - 「資格精算費」「資格費」「資格」→「資格精算費」
            - 判断不可→「その他経費」
        amount (int): 金額（円単位。1以上の正整数）
        expense_date (str): 経費発生日（YYYY-MM-DD形式）
        purpose (str): 業務目的（空文字禁止・500文字以内）
        tool_context: ツールコンテキスト（invocation_stateを含む）

    Returns:
        dict: 以下のキーを持つ辞書:
            - success (bool): 生成成否フラグ
            - file_path (str): 生成された申請書のファイルパス。エラー時は空文字列
            - application_data (dict): 申請書の各項目と記入値。エラー時は空辞書
            - message (str): エラーメッセージ（success=Trueの場合は空文字列）
    """
    # invocation_stateからapplicant_name・application_date・session_idを取得する
    applicant_name = ""
    application_date = ""
    session_id = ""
    file_path = ""

    if tool_context and hasattr(tool_context, "invocation_state") and tool_context.invocation_state:
        applicant_name = tool_context.invocation_state.get("applicant_name", "")
        application_date = tool_context.invocation_state.get("application_date", "")
        session_id = tool_context.invocation_state.get("session_id", "")

    logger.info(
        "generate_expense_application called: applicant=%s, date=%s",
        applicant_name, application_date,
    )

    try:
        # 申請者名の空文字チェック
        if not applicant_name or not applicant_name.strip():
            logger.error(
                "generate_expense_application validation error: applicant_name is empty, file_path=%s",
                file_path,
            )
            return {
                "success": False,
                "file_path": "",
                "application_data": {},
                "message": "入力のエラー: 申請者名が設定されていません。",
            }

        if not application_date or not application_date.strip():
            return {
                "success": False,
                "file_path": "",
                "application_data": {},
                "message": "日付のエラー: 申請日が正しく設定されていません。",
            }

        # ExpenseApplicationDataでバリデーション（申請期限チェックBRL-12含む）
        validated = ExpenseApplicationData(
            applicant_name=applicant_name,
            application_date=application_date,
            store_name=store_name,
            expense_category=expense_category,
            amount=amount,
            expense_date=expense_date,
            purpose=purpose,
        )

    except ValidationError as e:
        logger.warning(
            "generate_expense_application validation error: %s, file_path=%s", e, file_path
        )
        error_message = _error_handler.handle_validation_error(e)
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": error_message,
        }
    except Exception as e:
        logger.error(
            "generate_expense_application unexpected error: %s, file_path=%s", e, file_path,
            exc_info=True,
        )
        error_message = _error_handler.handle_unexpected_error(e)
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": error_message,
        }

    # 申請書IDとファイルパスを自律生成
    application_id = f"APP_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    logger.info("application_id generated: %s", application_id)

    file_name = f"{application_id}_経費精算申請.xlsx"
    output_dir = os.path.join(OUTPUT_BASE_DIR, session_id) if session_id else OUTPUT_BASE_DIR
    file_path = os.path.join(output_dir, file_name)

    # data/output/{session_id}/ ディレクトリを作成する
    try:
        _ensure_directory(output_dir)
    except Exception as e:
        logger.error(
            "generate_expense_application dir creation failed: %s, file_path=%s", e, file_path,
            exc_info=True,
        )
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": "出力ディレクトリの作成に失敗しました。担当部署へご連絡ください。",
        }

    # テンプレートファイルの存在チェック
    if not os.path.exists(DATA_EXPENSE_TEMPLATE_PATH):
        logger.error(
            "generate_expense_application template not found: %s, file_path=%s",
            DATA_EXPENSE_TEMPLATE_PATH, file_path,
        )
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": "申請書テンプレートの読み込みに失敗しました。担当部署へご連絡ください。",
        }

    # openpyxlでテンプレート読み込み・データマッピング
    try:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed")

        wb = openpyxl.load_workbook(DATA_EXPENSE_TEMPLATE_PATH)
        ws = wb.active

        # DD-01b 5.4節: 経費精算申請書
        # テンプレート構造: 1行目=タイトル, 3行目=申請者名ラベル(A3)/値(B3),
        #   4行目=申請日ラベル(A4)/値(B4), 6行目=ヘッダー, 7行目=明細, 9行目G=合計ラベル/H=合計値
        ws["B3"] = applicant_name
        ws["B4"] = application_date

        row = 7  # ヘッダー行(6)の次に明細を記入
        ws.cell(row=row, column=1, value=1)                              # A: No
        ws.cell(row=row, column=2, value=str(validated.expense_date))    # B: 購入日
        ws.cell(row=row, column=3, value=validated.store_name)           # C: 店舗名
        ws.cell(row=row, column=4, value=validated.expense_category)     # D: 品目
        ws.cell(row=row, column=5, value=validated.expense_category)     # E: 経費区分
        ws.cell(row=row, column=6, value=validated.amount)               # F: 金額
        ws.cell(row=row, column=7, value=validated.purpose)              # G: 業務目的
        ws.cell(row=row, column=8, value="")                             # H: 承認状況（空欄）

        ws["H9"] = validated.amount  # 合計金額

    except Exception as e:
        logger.error(
            "generate_expense_application template load error: %s, file_path=%s", e, file_path,
            exc_info=True,
        )
        error_message = _error_handler.handle_unexpected_error(e)
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": error_message,
        }

    # application_dataの構築
    application_data = {
        "application_id": application_id,
        "applicant_name": applicant_name,
        "application_date": application_date,
        "store_name": validated.store_name,
        "expense_category": validated.expense_category,
        "amount": validated.amount,
        "expense_date": str(validated.expense_date),
        "purpose": validated.purpose,
    }

    # Excelファイルを保存する（最大2回リトライ）
    save_result, save_error = _save_excel_file(wb, file_path)
    if not save_result:
        return {
            "success": False,
            "file_path": "",
            "application_data": {},
            "message": save_error,
        }

    # 監査ログ（DATA-008）に記録する
    _record_audit_log(
        application_id=application_id,
        applicant_name=applicant_name,
        application_type="経費精算申請",
        file_path=file_path,
    )

    logger.info("generate_expense_application succeeded: file_path=%s", file_path)
    return {
        "success": True,
        "file_path": file_path,
        "application_data": application_data,
        "message": "",
    }
