"""E2Eテスト: 申請書ファイルが実際に生成されることを確認する。

テスト内容:
1. 交通費精算申請書 (transport_expense_*.xlsx) の生成
2. 経費精算申請書 (expense_*.xlsx) の生成

いずれも実際のテンプレートファイルを読み込み、data/output/ 配下に出力する。
Strands エージェントへの API 呼び出しは行わず、ツール関数を直接呼び出す。
"""
import os
import sys
import glob
from unittest.mock import MagicMock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

PASS = "[PASS]"
FAIL = "[FAIL]"
results = []


def make_tool_context(session_id="e2e-test-session", applicant_name="山田太郎", application_date="2026-04-28"):
    ctx = MagicMock()
    ctx.invocation_state = {
        "session_id": session_id,
        "applicant_name": applicant_name,
        "application_date": application_date,
    }
    return ctx


# ── テスト 1: 運賃データ読み込み ────────────────────────────────────────
print("\n" + "=" * 60)
print("E2E テスト開始")
print("=" * 60)

print("\n[1] 運賃データ読み込みテスト")
try:
    from tools.transport_tools import _fare_loader
    ok_train, msg_train = _fare_loader.load_train_routes()
    assert ok_train, f"train_routes 読み込み失敗: {msg_train}"
    assert len(_fare_loader.train_routes_data) > 0, "train_routes データが空"
    print(f"  train_routes: {len(_fare_loader.train_routes_data)} 件ロード済み")

    ok_fixed, msg_fixed = _fare_loader.load_fixed_fares()
    assert ok_fixed, f"fixed_fares 読み込み失敗: {msg_fixed}"
    assert len(_fare_loader.fixed_fares_data) > 0, "fixed_fares データが空"
    print(f"  fixed_fares: {list(_fare_loader.fixed_fares_data.keys())}")
    print(f"  {PASS}")
    results.append(("運賃データ読み込み", True, ""))
except Exception as e:
    print(f"  {FAIL}: {e}")
    results.append(("運賃データ読み込み", False, str(e)))


# ── テスト 2: 電車区間運賃計算 ──────────────────────────────────────────
print("\n[2] 電車区間運賃計算テスト（渋谷→新宿）")
try:
    from tools.transport_tools import calculate_transport_expense
    ctx = make_tool_context()
    result = calculate_transport_expense(
        tool_context=ctx,
        transport_date="2026-04-28",
        departure="渋谷",
        destination="新宿",
        transport_type="電車",
    )
    assert result["success"] is True, f"計算失敗: {result.get('message')}"
    assert result["fare"] == 170, f"運賃が不正: {result['fare']}"
    print(f"  渋谷→新宿 電車: {result['fare']}円 （{result['calculation_basis']}）")
    print(f"  {PASS}")
    results.append(("電車区間運賃計算", True, ""))
except Exception as e:
    print(f"  {FAIL}: {e}")
    results.append(("電車区間運賃計算", False, str(e)))


# ── テスト 3: タクシー固定運賃計算 ─────────────────────────────────────
print("\n[3] タクシー固定運賃計算テスト")
try:
    from tools.transport_tools import calculate_transport_expense
    ctx = make_tool_context()
    result = calculate_transport_expense(
        tool_context=ctx,
        transport_date="2026-04-28",
        departure="東京",
        destination="品川",
        transport_type="タクシー",
    )
    assert result["success"] is True, f"計算失敗: {result.get('message')}"
    print(f"  タクシー固定運賃: {result['fare']}円")
    print(f"  {PASS}")
    results.append(("タクシー固定運賃計算", True, ""))
except Exception as e:
    print(f"  {FAIL}: {e}")
    results.append(("タクシー固定運賃計算", False, str(e)))


# ── テスト 4: 交通費精算申請書 Excel 生成 ───────────────────────────────
print("\n[4] 交通費精算申請書 Excel 生成テスト")
try:
    from tools.output_generator import generate_transport_expense_form
    session_id = "e2e-transport-20260428"
    ctx = make_tool_context(session_id=session_id)

    segments = [
        {
            "no": 1,
            "transport_date": "2026-04-25",
            "departure": "渋谷",
            "destination": "新宿",
            "transport_type": "電車",
            "amount": 170,
            "business_purpose": "取引先訪問",
        },
        {
            "no": 2,
            "transport_date": "2026-04-25",
            "departure": "新宿",
            "destination": "東京",
            "transport_type": "電車",
            "amount": 200,
            "business_purpose": "取引先訪問",
        },
    ]

    result = generate_transport_expense_form(
        tool_context=ctx,
        segments=segments,
        business_purpose="取引先訪問（営業部門）",
    )
    assert result["success"] is True, f"生成失敗: {result.get('message')}"
    file_path = result["file_path"]
    assert os.path.exists(file_path), f"ファイルが存在しない: {file_path}"
    size = os.path.getsize(file_path)
    assert size > 0, "ファイルサイズが0"
    print(f"  生成ファイル: {file_path}")
    print(f"  ファイルサイズ: {size:,} bytes")

    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    print(f"  申請者名セル(B3): {ws['B3'].value}")
    print(f"  申請日セル(B4):   {ws['B4'].value}")
    print(f"  1行目 出発地(C7): {ws['C7'].value}")
    print(f"  1行目 目的地(D7): {ws['D7'].value}")
    print(f"  1行目 金額(F7):   {ws['F7'].value}")
    print(f"  {PASS}")
    results.append(("交通費精算申請書生成", True, file_path))
except Exception as e:
    print(f"  {FAIL}: {e}")
    results.append(("交通費精算申請書生成", False, str(e)))


# ── テスト 5: 経費精算申請書 Excel 生成 ────────────────────────────────
print("\n[5] 経費精算申請書 Excel 生成テスト")
try:
    from tools.output_generator import generate_expense_form
    session_id = "e2e-expense-20260428"
    ctx = make_tool_context(session_id=session_id, applicant_name="田中花子")

    items = [
        {
            "no": 1,
            "purchase_date": "2026-04-22",
            "store_name": "文具屋アスクル",
            "item_name": "ボールペン（10本セット）",
            "expense_category": "事務用品費",
            "amount": 880,
            "business_purpose": "オフィス備品補充",
        },
        {
            "no": 2,
            "purchase_date": "2026-04-23",
            "store_name": "ビジネスホテル東京",
            "item_name": "宿泊費（1泊）",
            "expense_category": "宿泊費",
            "amount": 12000,
            "business_purpose": "大阪出張",
        },
    ]

    result = generate_expense_form(
        tool_context=ctx,
        items=items,
        business_purpose="出張・備品補充",
    )
    assert result["success"] is True, f"生成失敗: {result.get('message')}"
    file_path = result["file_path"]
    assert os.path.exists(file_path), f"ファイルが存在しない: {file_path}"
    size = os.path.getsize(file_path)
    assert size > 0, "ファイルサイズが0"
    print(f"  生成ファイル: {file_path}")
    print(f"  ファイルサイズ: {size:,} bytes")

    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    print(f"  申請者名セル(B3): {ws['B3'].value}")
    print(f"  申請日セル(B4):   {ws['B4'].value}")
    print(f"  1行目 品目(D7):   {ws['D7'].value}")
    print(f"  1行目 金額(F7):   {ws['F7'].value}")
    print(f"  {PASS}")
    results.append(("経費精算申請書生成", True, file_path))
except Exception as e:
    print(f"  {FAIL}: {e}")
    results.append(("経費精算申請書生成", False, str(e)))


# ── テスト 6: output ディレクトリの確認 ────────────────────────────────
print("\n[6] output ディレクトリ内ファイル一覧")
output_files = glob.glob("data/output/**/*.xlsx", recursive=True)
for f in sorted(output_files):
    size = os.path.getsize(f)
    print(f"  {f}  ({size:,} bytes)")


# ── 結果サマリ ─────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("テスト結果サマリ")
print("=" * 60)
passed = 0
failed = 0
for name, ok, detail in results:
    status = PASS if ok else FAIL
    print(f"  {status}  {name}")
    if ok and detail and detail.endswith(".xlsx"):
        print(f"         → {detail}")
    elif not ok:
        print(f"         → {detail}")
    if ok:
        passed += 1
    else:
        failed += 1

print(f"\n合計: {passed + failed} テスト / {passed} 成功 / {failed} 失敗")

if failed > 0:
    sys.exit(1)
